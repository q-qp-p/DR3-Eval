# Copyright 2025 Miromind.ai
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import argparse
import logging
import os
import re
import sys
from typing import Optional

from fastmcp import FastMCP
from mcp import ClientSession, StdioServerParameters
from mcp.client.stdio import stdio_client

logger = logging.getLogger("miroflow")

# Initialize FastMCP server
mcp = FastMCP("reading-mcp-server")

# Configuration: Strict limits to prevent context overload
MAX_PDF_PAGES_PER_READ = 3  # Only allow reading 3 pages at a time

# Try to import optional dependencies
try:
    import pdfminer.high_level
    from pdfminer.pdfpage import PDFPage
    HAS_PDFMINER = True
except ImportError:
    HAS_PDFMINER = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@mcp.tool()
async def read_pdf_pages(file_path: str, start_page: int = 1, end_page: Optional[int] = None) -> str:
    """Read specific pages from a PDF. LIMITED TO 3 PAGES MAX per call.
    
    ⚠️ IMPORTANT: Do NOT blindly read all pages! Follow this workflow:
    1. FIRST use search_in_file(file_path, keyword) to find which pages contain your target info
    2. The search results will tell you the exact page numbers
    3. THEN read only those specific pages (1-3 at a time)
    
    Args:
        file_path: The path to the PDF file.
        start_page: Starting page number (1-indexed). Default is 1.
        end_page: Ending page number. Auto-limited to start_page + 2 (max 3 pages).
    
    Returns:
        str: Text content of the specified pages.
    """
    if not HAS_PDFMINER:
        return "Error: pdfminer is not installed. Please install it with: pip install pdfminer.six"
    
    if not file_path or not file_path.strip():
        return "Error: file_path parameter is required and cannot be empty."
    
    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"
    
    if not file_path.lower().endswith('.pdf'):
        return "Error: File must be a PDF file."
    
    try:
        # Get total page count
        with open(file_path, 'rb') as f:
            pages = list(PDFPage.get_pages(f))
            total_pages = len(pages)
        
        # Validate page numbers
        if start_page < 1:
            start_page = 1
        if start_page > total_pages:
            return f"Error: start_page ({start_page}) exceeds total pages ({total_pages})."
        
        if end_page is None:
            end_page = start_page
        if end_page > total_pages:
            end_page = total_pages
        if end_page < start_page:
            end_page = start_page
        
        # Enforce strict page limit
        if end_page - start_page + 1 > MAX_PDF_PAGES_PER_READ:
            end_page = start_page + MAX_PDF_PAGES_PER_READ - 1
            if end_page > total_pages:
                end_page = total_pages
        
        # Extract text from specified pages (pdfminer uses 0-indexed pages)
        page_numbers = set(range(start_page - 1, end_page))
        text = pdfminer.high_level.extract_text(file_path, page_numbers=page_numbers)
        
        result = f"=== PDF: {os.path.basename(file_path)} ===\n"
        result += f"Pages {start_page}-{end_page} of {total_pages}\n\n"
        result += text
        
        return result
        
    except Exception as e:
        return f"Error reading PDF: {str(e)}"


@mcp.tool()
async def read_excel_rows(file_path: str, sheet_name: Optional[str] = None, start_row: int = 1, end_row: Optional[int] = None) -> str:
    """Read specific rows from an Excel file. Use this to read parts of a large Excel file without loading all data.
    
    Args:
        file_path: Required. The path to the Excel file (.xlsx or .xls).
        sheet_name: The name of the sheet to read. If not provided, reads the first sheet.
        start_row: The starting row number (1-indexed). Default is 1.
        end_row: The ending row number (1-indexed, inclusive). If not provided, reads up to 100 rows from start_row.
    
    Returns:
        str: The content of the specified rows in markdown table format, or an error message if reading fails.
    """
    if not HAS_OPENPYXL:
        return "Error: openpyxl is not installed. Please install it with: pip install openpyxl"
    
    if not file_path or not file_path.strip():
        return "Error: file_path parameter is required and cannot be empty."
    
    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"
    
    if not file_path.lower().endswith(('.xlsx', '.xls')):
        return "Error: File must be an Excel file (.xlsx or .xls)."
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Get sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return f"Error: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
            sheet_name = sheet.title
        
        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Validate row numbers
        if start_row < 1:
            start_row = 1
        if start_row > max_row:
            return f"Error: start_row ({start_row}) exceeds total rows ({max_row})."
        
        if end_row is None:
            end_row = min(start_row + 99, max_row)  # Default to 100 rows
        if end_row > max_row:
            end_row = max_row
        if end_row < start_row:
            end_row = start_row
        
        # Build markdown table
        result = f"=== Excel: {os.path.basename(file_path)} ===\n"
        result += f"Sheet: {sheet_name}\n"
        result += f"Rows {start_row}-{end_row} of {max_row}, Columns: {max_col}\n\n"
        
        # Header row
        result += "|"
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=start_row, column=col_idx)
            cell_value = str(cell.value) if cell.value is not None else ""
            result += f" {cell_value} |"
        result += "\n"
        
        # Separator
        result += "|"
        for _ in range(max_col):
            result += " --- |"
        result += "\n"
        
        # Data rows
        for row_idx in range(start_row + 1, end_row + 1):
            result += "|"
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_value = str(cell.value) if cell.value is not None else ""
                result += f" {cell_value} |"
            result += "\n"
        
        return result
        
    except Exception as e:
        return f"Error reading Excel: {str(e)}"


@mcp.tool()
async def search_in_file(file_path: str, keyword: str, context_chars: int = 500) -> str:
    """Search for a keyword in a file. Returns page numbers AND detailed context.
    
    ⚠️ ALWAYS use this FIRST! The context returned is usually enough to answer your question.
    Only use read_pdf_pages if you need more details from a specific page.
    
    Args:
        file_path: The path to the file to search.
        keyword: The keyword or phrase to search for (case-insensitive).
        context_chars: Characters of context around each match. Default is 500.
    
    Returns:
        str: Page numbers with detailed context. Often sufficient without reading full pages.
    """
    if not file_path or not file_path.strip():
        return "Error: file_path parameter is required and cannot be empty."
    
    if not keyword or not keyword.strip():
        return "Error: keyword parameter is required and cannot be empty."
    
    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"
    
    ext = os.path.splitext(file_path)[1].lower()
    
    try:
        # Special handling for PDF: search page by page to get page numbers
        if ext == '.pdf':
            if not HAS_PDFMINER:
                return "Error: pdfminer is not installed for PDF search."
            
            # Get total pages
            with open(file_path, 'rb') as f:
                pdf_pages = list(PDFPage.get_pages(f))
                total_pages = len(pdf_pages)
            
            # Search each page
            pattern = re.compile(re.escape(keyword), re.IGNORECASE)
            page_matches = []  # List of (page_num, match_count, first_context)
            
            for page_num in range(total_pages):
                page_text = pdfminer.high_level.extract_text(file_path, page_numbers={page_num})
                matches = list(pattern.finditer(page_text))
                if matches:
                    # Get context around first match
                    first_match = matches[0]
                    start = max(0, first_match.start() - context_chars)
                    end = min(len(page_text), first_match.end() + context_chars)
                    context = page_text[start:end].strip()
                    # Highlight keyword
                    context = pattern.sub(f"**{first_match.group()}**", context)
                    page_matches.append((page_num + 1, len(matches), context))
            
            if not page_matches:
                return f"No matches found for '{keyword}' in {os.path.basename(file_path)} ({total_pages} pages)"
            
            result = f"=== Search Results: '{keyword}' in {os.path.basename(file_path)} ===\n"
            result += f"Total pages: {total_pages}\n"
            result += f"Found matches on {len(page_matches)} page(s)\n\n"
            
            # List pages with matches
            result += "📍 PAGES WITH MATCHES:\n"
            for page_num, count, _ in page_matches:
                result += f"  - Page {page_num}: {count} match(es)\n"
            
            result += "\n💡 To read these pages, use: read_pdf_pages(file_path, page_number)\n\n"
            
            # Show detailed context for matches (this is often enough to answer questions)
            result += "--- Detailed Context (may contain your answer) ---\n"
            for page_num, count, context in page_matches[:3]:
                result += f"\n[Page {page_num}] ({count} matches):\n"
                result += f"{context}\n"
            
            if len(page_matches) > 3:
                result += f"\n... {len(page_matches) - 3} more pages have matches. "
                result += "Only read those pages if the above context doesn't answer your question.\n"
            
            return result
        
        # For Excel files
        elif ext in ['.xlsx', '.xls']:
            if not HAS_OPENPYXL:
                return "Error: openpyxl is not installed for Excel search."
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            pattern = re.compile(re.escape(keyword), re.IGNORECASE)
            row_matches = []  # List of (sheet_name, row_num, context)
            
            for sheet in wb.worksheets:
                for row_idx, row in enumerate(sheet.iter_rows(), 1):
                    row_text = " | ".join(str(cell.value) if cell.value else "" for cell in row)
                    if pattern.search(row_text):
                        highlighted = pattern.sub(f"**\\g<0>**", row_text)
                        row_matches.append((sheet.title, row_idx, highlighted[:200]))
            
            if not row_matches:
                return f"No matches found for '{keyword}' in {os.path.basename(file_path)}"
            
            result = f"=== Search Results: '{keyword}' in {os.path.basename(file_path)} ===\n"
            result += f"Found matches in {len(row_matches)} row(s)\n\n"
            
            result += "📍 ROWS WITH MATCHES:\n"
            for sheet_name, row_num, context in row_matches[:10]:
                result += f"  - Sheet '{sheet_name}', Row {row_num}: {context}\n"
            
            if len(row_matches) > 10:
                result += f"\n... and {len(row_matches) - 10} more rows.\n"
            
            result += "\n💡 To read specific rows, use: read_excel_rows(file_path, sheet_name, start_row, end_row)\n"
            
            return result
        
        # For text files
        elif ext in ['.txt', '.md', '.csv', '.json']:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            pattern = re.compile(re.escape(keyword), re.IGNORECASE)
            line_matches = []  # List of (line_num, context)
            
            for line_num, line in enumerate(lines, 1):
                if pattern.search(line):
                    highlighted = pattern.sub(f"**\\g<0>**", line.strip())
                    line_matches.append((line_num, highlighted[:200]))
            
            if not line_matches:
                return f"No matches found for '{keyword}' in {os.path.basename(file_path)}"
            
            result = f"=== Search Results: '{keyword}' in {os.path.basename(file_path)} ===\n"
            result += f"Found matches on {len(line_matches)} line(s)\n\n"
            
            result += "📍 LINES WITH MATCHES:\n"
            for line_num, context in line_matches[:15]:
                result += f"  Line {line_num}: {context}\n"
            
            if len(line_matches) > 15:
                result += f"\n... and {len(line_matches) - 15} more lines.\n"
            
            return result
        
        else:
            return f"Error: Unsupported file type: {ext}. Supported: .pdf, .xlsx, .xls, .txt, .md, .csv, .json"
        
    except Exception as e:
        return f"Error searching file: {str(e)}"


@mcp.tool()
async def get_file_info(file_path: str) -> str:
    """Get information about a file without reading its full content.
    Useful for understanding file structure before reading specific parts.
    
    Args:
        file_path: Required. The path to the file.
    
    Returns:
        str: File information including size, type, and structure details.
    """
    if not file_path or not file_path.strip():
        return "Error: file_path parameter is required and cannot be empty."
    
    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"
    
    ext = os.path.splitext(file_path)[1].lower()
    size_bytes = os.path.getsize(file_path)
    size_kb = size_bytes / 1024
    size_mb = size_kb / 1024
    
    result = f"=== File Info: {os.path.basename(file_path)} ===\n"
    result += f"Path: {file_path}\n"
    result += f"Size: {size_mb:.2f} MB ({size_kb:.1f} KB)\n"
    result += f"Type: {ext}\n\n"
    
    try:
        if ext == '.pdf':
            if HAS_PDFMINER:
                with open(file_path, 'rb') as f:
                    pages = list(PDFPage.get_pages(f))
                    result += f"Total Pages: {len(pages)}\n"
                    result += "\nTo read specific pages, use: read_pdf_pages(file_path, start_page, end_page)\n"
            else:
                result += "Note: Install pdfminer.six for detailed PDF info.\n"
                
        elif ext in ['.xlsx', '.xls']:
            if HAS_OPENPYXL:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                result += f"Sheets: {wb.sheetnames}\n"
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    result += f"  - {sheet_name}: {sheet.max_row} rows x {sheet.max_column} columns\n"
                result += "\nTo read specific rows, use: read_excel_rows(file_path, sheet_name, start_row, end_row)\n"
            else:
                result += "Note: Install openpyxl for detailed Excel info.\n"
                
        elif ext in ['.txt', '.md', '.csv', '.json']:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
                result += f"Total Lines: {len(lines)}\n"
                # Show first few lines as preview
                result += "\nPreview (first 10 lines):\n"
                for i, line in enumerate(lines[:10]):
                    result += f"  {i+1}: {line.rstrip()[:100]}\n"
                if len(lines) > 10:
                    result += f"  ... and {len(lines) - 10} more lines\n"
        else:
            result += "Note: Detailed info not available for this file type.\n"
            
    except Exception as e:
        result += f"Error getting details: {str(e)}\n"
    
    return result


@mcp.tool()
async def convert_to_markdown(uri: str) -> str:
    """Convert various types of resources (doc, ppt, pdf, excel, csv, zip file etc.)
    described by an file: or data: URI to markdown.

    Args:
        uri: Required. The URI of the resource to convert. Need to start with 'file:' or 'data:' schemes.

    Returns:
        str: The converted markdown content, or an error message if conversion fails.
    """
    if not uri or not uri.strip():
        return "Error: URI parameter is required and cannot be empty."

    # Validate URI scheme
    valid_schemes = ["http:", "https:", "file:", "data:"]
    if not any(uri.lower().startswith(scheme) for scheme in valid_schemes):
        return f"Error: Invalid URI scheme. Supported schemes are: {', '.join(valid_schemes)}"

    tool_name = "convert_to_markdown"
    arguments = {"uri": uri}

    server_params = StdioServerParameters(
        command=sys.executable,
        args=["-m", "markitdown_mcp"],
    )

    result_content = ""
    try:
        async with stdio_client(server_params) as (read, write):
            async with ClientSession(read, write, sampling_callback=None) as session:
                await session.initialize()
                try:
                    tool_result = await session.call_tool(
                        tool_name, arguments=arguments
                    )
                    result_content = (
                        tool_result.content[-1].text if tool_result.content else ""
                    )
                except Exception as tool_error:
                    logger.info(f"Tool execution error: {tool_error}")
                    return f"Error: Tool execution failed: {str(tool_error)}"
    except Exception as session_error:
        logger.info(f"Session error: {session_error}")
        return (
            f"Error: Failed to connect to markitdown-mcp server: {str(session_error)}"
        )

    return result_content


if __name__ == "__main__":
    # Set up argument parser
    parser = argparse.ArgumentParser(description="Reading MCP Server")
    parser.add_argument(
        "--transport",
        choices=["stdio", "http"],
        default="stdio",
        help="Transport method: 'stdio' or 'http' (default: stdio)",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=8080,
        help="Port to use when running with HTTP transport (default: 8080)",
    )
    parser.add_argument(
        "--path",
        type=str,
        default="/mcp",
        help="URL path to use when running with HTTP transport (default: /mcp)",
    )

    # Parse command line arguments
    args = parser.parse_args()

    # Run the server with the specified transport method
    if args.transport == "stdio":
        mcp.run(transport="stdio")
    else:
        # For HTTP transport, include port and path options
        mcp.run(transport="streamable-http", port=args.port, path=args.path)
